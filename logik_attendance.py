import tkinter as tk
from tkinter import ttk
from ttkthemes import ThemedTk
import sqlite3
import os
import pandas as pd
from tkinter import filedialog
from tkinter.messagebox import OK, INFO, showinfo
from tkcalendar import Calendar 

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



def create_attendance_stud(da, nd, m, id_st):
    conn=sqlite3.connect("university.db", check_same_thread=False)
    cur=conn.cursor()
    try:
        cur.execute("INSERT INTO attendance (date_attendance, name_discipline, mark, id_student) VALUES (?,?,?,?)",(da.selection_get(), nd.get(), m, id_st))
    except Exception as e:
         print(e)
    conn.commit()
    
def load_att_stud(k, f, n, d, trv):
    print('qwerty')
    for item in trv.get_children():
        trv.delete(item)
            
    conn=sqlite3.connect("university.db", check_same_thread=False)
    cur=conn.cursor()
    cur.execute("""select id_attendance, student.id_student, 
                surname, name, oldname, strftime('%d.%m.%Y', date_attendance) AS formatted_date,
                 mark from attendance left join student
                 on student.id_student = attendance.id_student
                left JOIN student_group on student_group.id_student = student.id_student
                 left join university_group on
                 university_group.id_univ_group = student_group.id_univ_group
                 where course=? AND faculty=? AND number=? AND name_discipline=?""",
                   (k.get(), f.get(), n.get(), d.get()))
    rows = cur.fetchall()
        
    for row in rows:
        trv.insert("", tk.END, values=row)

def upload_result(k, f, n, d):
    conn = sqlite3.connect('university.db') 
    sql_query =f"""
    select  
                surname || ' ' || name || ' ' || oldname AS ФИО_студента,
                strftime('%d.%m.%Y', date_attendance) AS formatted_date,
                 mark from attendance left join student
                 on student.id_student = attendance.id_student
                left JOIN student_group on student_group.id_student = student.id_student
                 left join university_group on
                 university_group.id_univ_group = student_group.id_univ_group
                 where course='{k.get()}' AND faculty='{f.get()}' AND number='{n.get()}' AND name_discipline='{d.get()}'
    """
    
    df = pd.read_sql_query(sql_query, conn)
    conn.close()

    pivot_table = df.pivot_table(
        index='ФИО_студента',
        columns='formatted_date',
        values='mark',
        aggfunc='first',            
        fill_value=''
    )

    print(pivot_table)

    pivot_table.to_excel("Посещаемость-" + k.get() + "-" + f.get() + "-" + n.get() + " " + d.get() +  ".xlsx")
    
    file_path = "Посещаемость-" + k.get() + "-" + f.get() + "-" + n.get() + " " + d.get() +  ".xlsx"
    wb = load_workbook(file_path)
    ws = wb.active
    ws.column_dimensions['A'].width = 40
                    
    for i in range(2, ws.max_column + 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 15
                    
    ws.freeze_panes = 'B1'
    wb.save(file_path)
    wb.close()